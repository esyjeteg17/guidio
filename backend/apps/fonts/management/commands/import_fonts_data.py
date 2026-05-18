from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from apps.fonts.models import Font, FontPair


DATA_DIR = Path(settings.BASE_DIR) / "apps" / "fonts" / "data"
FONTS_XLSX = DATA_DIR / "fonts.xlsx"
PAIRS_XLSX = DATA_DIR / "pairs.xlsx"


def detect_source(link: str) -> tuple[str, str]:
    """Returns (source, google_family) based on the raw link cell."""
    link = (link or "").strip()
    if "fonts.google.com" in link:
        parsed = urlparse(link)
        parts = [p for p in parsed.path.split("/") if p]
        family = ""
        if "specimen" in parts:
            idx = parts.index("specimen")
            if idx + 1 < len(parts):
                family = parts[idx + 1].replace("+", " ")
        return Font.Source.GOOGLE_FONTS, family
    if link.lower() == "локально":
        return Font.Source.LOCAL, ""
    return Font.Source.EXTERNAL, ""


class Command(BaseCommand):
    help = "Импортирует шрифты и шрифтовые пары из xlsx-таблиц"

    def add_arguments(self, parser):
        parser.add_argument(
            "--wipe",
            action="store_true",
            help="Удалить все существующие Font/FontPair перед импортом",
        )

    @transaction.atomic
    def handle(self, *args, wipe: bool = False, **options):
        if wipe:
            FontPair.objects.all().delete()
            Font.objects.all().delete()
            self.stdout.write(self.style.WARNING("Existing fonts and pairs wiped."))

        created_fonts, skipped_pairs = self._import_fonts(), []
        self.stdout.write(
            self.style.SUCCESS(f"Шрифтов в БД: {Font.objects.count()} (создано/обновлено: {created_fonts})")
        )

        pair_count, skipped_pairs = self._import_pairs()
        self.stdout.write(
            self.style.SUCCESS(f"Пар в БД: {FontPair.objects.count()} (создано/обновлено: {pair_count})")
        )
        if skipped_pairs:
            self.stdout.write(
                self.style.WARNING(
                    f"Пропущено пар (шрифт не найден): {len(skipped_pairs)}"
                )
            )
            for row, heading, body in skipped_pairs:
                self.stdout.write(f"  row {row}: '{heading}' + '{body}'")

    def _import_fonts(self) -> int:
        wb = openpyxl.load_workbook(FONTS_XLSX, data_only=True)
        ws = wb["Шрифты"]
        count = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row[0] is None or not str(row[0]).strip():
                continue
            name = str(row[0]).strip()
            category = (row[1] or "").strip()
            role = (row[2] or "").strip()
            usage_context = (row[3] or "").strip()
            theme = (row[4] or "").strip()
            character_mood = (row[5] or "").strip()
            style = (row[6] or "").strip()
            width = (row[7] or "").strip()
            link = (row[8] or "").strip()
            embedding_text = (row[9] or "").strip()

            source, google_family = detect_source(link)

            Font.objects.update_or_create(
                name=name,
                defaults=dict(
                    category=category,
                    role=role,
                    width=width,
                    usage_context=usage_context,
                    theme=theme,
                    character_mood=character_mood,
                    style=style,
                    link=link,
                    source=source,
                    google_family=google_family,
                    embedding_text=embedding_text,
                ),
            )
            count += 1
        return count

    def _import_pairs(self) -> tuple[int, list]:
        wb = openpyxl.load_workbook(PAIRS_XLSX, data_only=True)
        ws = wb["Пары"]
        count = 0
        skipped: list[tuple[int, str, str]] = []
        font_index = {f.name: f for f in Font.objects.all()}

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if row[0] is None or row[1] is None:
                continue
            heading_name = str(row[0]).strip()
            body_name = str(row[1]).strip()
            context = (row[2] or "").strip()

            heading = font_index.get(heading_name)
            body = font_index.get(body_name)
            if not heading or not body:
                skipped.append((i, heading_name, body_name))
                continue

            FontPair.objects.update_or_create(
                heading=heading,
                body=body,
                defaults=dict(context=context),
            )
            count += 1
        return count, skipped
